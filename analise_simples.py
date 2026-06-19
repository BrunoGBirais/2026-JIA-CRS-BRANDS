import openpyxl
from openpyxl import load_workbook

arquivo = r'c:\Users\bruno.birais\Documents\CRS-Brants\6- Cenário de Demanda 15.06.2026.xlsx'

# Carregar o workbook sem carregar dados
wb = load_workbook(arquivo, data_only=False)
ws = wb['Metodologia']

print("ESTRUTURA DA METODOLOGIA")
print("=" * 80)

# Headers (linha 5)
print("\nCOLUNAS PRINCIPAIS (11-30):")
for col in range(11, 31):
    header = ws.cell(5, col).value
    print(f"  Col {col}: {header}")

print("\n\nFÓRMULAS NA LINHA 6 (Item 1211, R01) - Colunas 11 a 30:")
print("=" * 80)
for col in range(11, 31):
    cell = ws.cell(6, col)
    header = ws.cell(5, col).value
    
    if cell.data_type == 'f':
        print(f"\nCol {col} ({header}): FÓRMULA")
        print(f"  {cell.value}")
    elif cell.value is not None:
        print(f"\nCol {col} ({header}): VALOR = {cell.value}")

print("\n\nFÓRMULAS NAS ÚLTIMAS 15 COLUNAS (Forecast futuro):")
print("=" * 80)
max_col = ws.max_column
print(f"Total de colunas: {max_col}")

for col in range(max_col - 14, max_col + 1):
    cell = ws.cell(6, col)
    header = ws.cell(5, col).value
    
    if cell.data_type == 'f':
        formula_resumida = cell.value[:100] if len(str(cell.value)) > 100 else cell.value
        print(f"\nCol {col} ({header}): FÓRMULA")
        print(f"  {formula_resumida}...")
    elif cell.value is not None:
        print(f"\nCol {col} ({header}): VALOR = {cell.value}")

wb.close()
print("\n\nFINALIZADO!")
