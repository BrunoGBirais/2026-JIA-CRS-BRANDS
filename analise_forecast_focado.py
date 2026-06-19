import openpyxl
from openpyxl import load_workbook

arquivo = r'c:\Users\bruno.birais\Documents\CRS-Brants\6- Cenário de Demanda 15.06.2026.xlsx'

wb = load_workbook(arquivo, data_only=False)
ws = wb['Metodologia']

# Função para converter número para letra da coluna
def col_to_letter(n):
    result = ""
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        result = chr(65 + remainder) + result
    return result

print("=" * 100)
print("ANÁLISE FOCADA: CÁLCULO DOS 12 MESES DE FORECAST")
print("=" * 100)

linha = 6  # Exemplo: Item 1211, R01

# Analisar as colunas 140-151 (12 meses de forecast)
print("\n" + "=" * 100)
print("12 MESES DE FORECAST (Colunas 140-151):")
print("=" * 100)

for col in range(140, 152):
    letra = col_to_letter(col)
    header = ws.cell(5, col).value
    cell = ws.cell(linha, col)
    
    print(f"\nMÊS {col-139} - Col {col} ({letra}) - Header: {header}")
    
    if cell.data_type == 'f':
        print(f"  FÓRMULA: {cell.value}")
    else:
        print(f"  VALOR: {cell.value}")

# Agora mapear as colunas BC (55) e CQ (95) para entender o padrão
print("\n\n" + "=" * 100)
print("COLUNAS BC a BN (55-66) - Primeira parte da subtração:")
print("=" * 100)

for col in range(55, 67):
    letra = col_to_letter(col)
    header = ws.cell(5, col).value
    cell = ws.cell(linha, col)
    
    if cell.data_type == 'f':
        print(f"Col {col} ({letra}): {header} = FÓRMULA: {cell.value[:80]}")
    else:
        print(f"Col {col} ({letra}): {header} = VALOR: {cell.value}")

print("\n\n" + "=" * 100)
print("COLUNAS CQ a DB (95-106) - Segunda parte da subtração:")
print("=" * 100)

for col in range(95, 107):
    letra = col_to_letter(col)
    header = ws.cell(5, col).value
    cell = ws.cell(linha, col)
    
    if cell.data_type == 'f':
        print(f"Col {col} ({letra}): {header} = FÓRMULA: {cell.value[:80]}")
    else:
        print(f"Col {col} ({letra}): {header} = VALOR: {cell.value}")

# Verificar o que tem nas linhas 1-3 para entender os anos
print("\n\n" + "=" * 100)
print("IDENTIFICANDO OS ANOS (linhas 1-3):")
print("=" * 100)

anos_importantes = [11, 25, 39, 55, 95, 140]  # Colunas chave
for row in range(1, 4):
    print(f"\nLinha {row}:")
    for col in anos_importantes:
        letra = col_to_letter(col)
        valor = ws.cell(row, col).value
        if valor:
            print(f"  Col {col} ({letra}): {valor}")

wb.close()
print("\n\nCONCLUÍDO!")
