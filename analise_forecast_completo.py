import openpyxl
from openpyxl import load_workbook

arquivo = r'c:\Users\bruno.birais\Documents\CRS-Brants\6- Cenário de Demanda 15.06.2026.xlsx'

wb = load_workbook(arquivo, data_only=False)
ws = wb['Metodologia']

print("=" * 100)
print("MAPEAMENTO COMPLETO DAS COLUNAS - ABA METODOLOGIA")
print("=" * 100)

max_col = ws.max_column
print(f"\nTotal de colunas: {max_col}")

# Função para converter número de coluna para letra
def col_num_to_letter(n):
    result = ""
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        result = chr(65 + remainder) + result
    return result

# Mapear TODAS as colunas (headers da linha 5)
print("\n" + "=" * 100)
print("TODAS AS COLUNAS (Headers da linha 5):")
print("=" * 100)

colunas_importantes = {}

for col_num in range(1, max_col + 1):
    header = ws.cell(5, col_num).value
    letra = col_num_to_letter(col_num)
    
    # Guardar colunas importantes para referência
    if header:
        colunas_importantes[letra] = header
    
    # Mostrar todas as colunas
    print(f"Col {col_num:3d} ({letra:>3s}): {header}")

# Agora analisar especificamente as colunas usadas no forecast (140-151)
print("\n" + "=" * 100)
print("ANÁLISE DETALHADA DOS 12 MESES DE FORECAST (Colunas 140-151)")
print("=" * 100)

linha_exemplo = 6  # Item 1211, R01

for col in range(140, 152):
    letra = col_num_to_letter(col)
    header = ws.cell(5, col).value
    cell = ws.cell(linha_exemplo, col)
    
    print(f"\n{'='*80}")
    print(f"COLUNA {col} ({letra}) - Header: {header}")
    print(f"{'='*80}")
    
    if cell.data_type == 'f':
        formula = cell.value
        print(f"FÓRMULA: {formula}")
        
        # Tentar identificar quais colunas são referenciadas
        import re
        # Extrair referências de células (ex: CQ6, BC6)
        refs = re.findall(r'[A-Z]+\d+', formula)
        
        if refs:
            print(f"\nReferências encontradas:")
            for ref in refs:
                # Extrair letra da coluna
                col_ref = re.findall(r'[A-Z]+', ref)[0]
                # Tentar buscar o que é essa coluna
                # Converter letra para número
                col_num_ref = sum((ord(c) - 64) * (26 ** i) for i, c in enumerate(reversed(col_ref)))
                header_ref = ws.cell(5, col_num_ref).value
                print(f"  {ref} (Col {col_num_ref}, {col_ref}): {header_ref}")
    else:
        print(f"VALOR: {cell.value}")

# Identificar o padrão das colunas BC, BD, BE... e CQ, CR, CS...
print("\n" + "=" * 100)
print("MAPEAMENTO DAS COLUNAS REFERENCIADAS NO FORECAST:")
print("=" * 100)

# Colunas BC em diante (provavelmente orçamento)
print("\nColunas BC a BN (usadas como subtração):")
for col in range(55, 67):  # BC=55, BN=66
    letra = col_num_to_letter(col)
    header = ws.cell(5, col).value
    valor_ex = ws.cell(linha_exemplo, col).value
    print(f"  Col {col:3d} ({letra}): {header} = {valor_ex}")

# Colunas CQ em diante (provavelmente forecast base)
print("\nColunas CQ a DB (usadas como base):")
for col in range(95, 107):  # CQ=95
    letra = col_num_to_letter(col)
    header = ws.cell(5, col).value
    cell = ws.cell(linha_exemplo, col)
    
    if cell.data_type == 'f':
        print(f"  Col {col:3d} ({letra}): {header} = FÓRMULA: {cell.value[:60]}...")
    else:
        print(f"  Col {col:3d} ({letra}): {header} = {cell.value}")

wb.close()
print("\n\nANÁLISE FINALIZADA!")
