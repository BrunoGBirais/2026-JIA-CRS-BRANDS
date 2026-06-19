import openpyxl
from openpyxl import load_workbook

arquivo = r'c:\Users\bruno.birais\Documents\CRS-Brants\6- Cenário de Demanda 15.06.2026.xlsx'

print("=" * 100)
print("ANÁLISE DAS FÓRMULAS E METODOLOGIA DE FORECAST")
print("=" * 100)

# Carregar o workbook
wb = load_workbook(arquivo, data_only=False)

# Analisar a aba Metodologia
ws_metodologia = wb['Metodologia']

print("\n" + "=" * 100)
print("ESTRUTURA DA ABA 'METODOLOGIA':")
print("=" * 100)

# Pegar os headers da linha 5 (índice 5)
headers = []
for col in range(1, 30):
    cell = ws_metodologia.cell(row=5, column=col)
    headers.append(f"Col {col}: {cell.value}")
    
for h in headers:
    print(h)

# Analisar as fórmulas em uma linha de exemplo (linha 6 - primeiro produto)
print("\n" + "=" * 100)
print("ANÁLISE DAS FÓRMULAS - LINHA 6 (Primeiro Produto - Item 1211, R01):")
print("=" * 100)

linha_exemplo = 6
for col in range(11, 25):  # Colunas após os dados básicos
    cell = ws_metodologia.cell(row=linha_exemplo, column=col)
    header = ws_metodologia.cell(row=5, column=col).value
    
    if cell.data_type == 'f':  # É fórmula
        print(f"\nCol {col} ({header}):")
        print(f"  Fórmula: {cell.value}")
    else:
        print(f"\nCol {col} ({header}): Valor = {cell.value}")

# Verificar colunas mais à frente (forecast futuro)
print("\n" + "=" * 100)
print("ANÁLISE DAS COLUNAS DE FORECAST (últimas 20 colunas):")
print("=" * 100)

max_col = ws_metodologia.max_column
for col in range(max_col - 20, max_col + 1):
    cell_header = ws_metodologia.cell(row=5, column=col)
    cell_data = ws_metodologia.cell(row=linha_exemplo, column=col)
    
    print(f"\nCol {col} - Header: {cell_header.value}")
    if cell_data.data_type == 'f':
        print(f"  Fórmula: {cell_data.value}")
    else:
        print(f"  Valor: {cell_data.value}")

# Analisar linha 4 (linha de exemplo/totais)
print("\n" + "=" * 100)
print("ANÁLISE DA LINHA 4 (EXEMPLO/TOTAIS):")
print("=" * 100)

for col in range(11, 25):
    cell = ws_metodologia.cell(row=4, column=col)
    header = ws_metodologia.cell(row=5, column=col).value
    
    print(f"\nCol {col} ({header}):")
    if cell.data_type == 'f':
        print(f"  Fórmula: {cell.value}")
    else:
        print(f"  Valor: {cell.value}")

# Verificar se há texto explicativo em células mescladas ou escondidas
print("\n" + "=" * 100)
print("PROCURANDO TEXTO EXPLICATIVO (primeiras 3 linhas, todas as colunas):")
print("=" * 100)

for row in range(1, 4):
    for col in range(1, 20):
        cell = ws_metodologia.cell(row=row, column=col)
        if cell.value and str(cell.value).strip():
            print(f"Linha {row}, Col {col}: {cell.value}")

wb.close()
